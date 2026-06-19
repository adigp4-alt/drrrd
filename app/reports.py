"""Report generation for Excel export, PDF export, and performance summaries."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.config import TIERS


def generate_excel_report(holdings, current_prices):
    """Generate an Excel report with portfolio data and charts."""
    wb = Workbook()

    # --- Portfolio Holdings Sheet ---
    ws = wb.active
    ws.title = "Portfolio"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Ticker", "Name", "Tier", "Shares", "Buy Price", "Current Price",
               "Cost Basis", "Market Value", "P&L ($)", "P&L (%)", "Allocation %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    total_value = 0
    rows_data = []
    for h in holdings:
        ticker = h["ticker"]
        price_data = current_prices.get(ticker, {})
        current = price_data.get("price", h["buy_price"])
        cost_basis = h["shares"] * h["buy_price"]
        market_value = h["shares"] * current
        pnl = market_value - cost_basis
        pnl_pct = (pnl / cost_basis * 100) if cost_basis else 0
        total_value += market_value
        rows_data.append({
            "ticker": ticker,
            "name": price_data.get("name", ticker),
            "tier": price_data.get("tier", ""),
            "shares": h["shares"],
            "buy_price": h["buy_price"],
            "current": current,
            "cost_basis": cost_basis,
            "market_value": market_value,
            "pnl": pnl,
            "pnl_pct": pnl_pct,
        })

    for i, rd in enumerate(rows_data):
        row = i + 2
        alloc = (rd["market_value"] / total_value * 100) if total_value else 0
        values = [rd["ticker"], rd["name"], rd["tier"], rd["shares"],
                  rd["buy_price"], rd["current"], rd["cost_basis"],
                  rd["market_value"], rd["pnl"], rd["pnl_pct"], alloc]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=round(val, 2) if isinstance(val, float) else val)
            cell.border = thin_border
            if col in (5, 6, 7, 8, 9):
                cell.number_format = '$#,##0.00'
            elif col in (10, 11):
                cell.number_format = '0.00"%"'

    # P&L chart
    if rows_data:
        chart = BarChart()
        chart.title = "P&L by Position"
        chart.y_axis.title = "P&L ($)"
        data = Reference(ws, min_col=9, min_row=1, max_row=len(rows_data) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows_data) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 20
        ws.add_chart(chart, "A" + str(len(rows_data) + 4))

    # --- Market Overview Sheet ---
    ws2 = wb.create_sheet("Market Overview")
    headers2 = ["Ticker", "Name", "Tier", "Price", "Change %", "Volume"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for ticker, data in current_prices.items():
        values = [ticker, data.get("name", ""), data.get("tier", ""),
                  data.get("price", 0), data.get("change_pct", 0), data.get("volume", 0)]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border
        row += 1

    # Auto-size columns
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_performance_summary(current_prices, history_data):
    """Generate a performance summary dict."""
    if not current_prices:
        return {"error": "No data available"}

    tickers = list(current_prices.keys())
    sorted_by_change = sorted(tickers, key=lambda t: current_prices[t].get("change_pct", 0))

    top_gainers = []
    for t in reversed(sorted_by_change[-5:]):
        d = current_prices[t]
        top_gainers.append({"ticker": t, "change_pct": d["change_pct"], "price": d["price"]})

    top_losers = []
    for t in sorted_by_change[:5]:
        d = current_prices[t]
        top_losers.append({"ticker": t, "change_pct": d["change_pct"], "price": d["price"]})

    # Tier performance
    tier_perf = {}
    for tid, tdata in TIERS.items():
        changes = []
        for sym in tdata["tickers"]:
            if sym in current_prices:
                changes.append(current_prices[sym].get("change_pct", 0))
        if changes:
            tier_perf[tid] = {
                "name": tdata["name"],
                "avg_change": round(sum(changes) / len(changes), 2),
                "best": max(changes),
                "worst": min(changes),
            }

    return {
        "total_tickers": len(current_prices),
        "top_gainers": top_gainers,
        "top_losers": top_losers,
        "tier_performance": tier_perf,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _styled_table(data, header_color, col_widths=None):
    """Build a reportlab Table with consistent styling."""
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_color)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    return table


def generate_pdf_report(holdings, current_prices):
    """Generate a PDF report with market overview, gainers/losers, and holdings."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter),
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Title'],
                                 fontSize=18, spaceAfter=12)
    elements = [
        Paragraph("Iran Investment Tracker — Report", title_style),
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']),
        Spacer(1, 20),
    ]

    # Market Overview (sorted by change desc)
    elements.append(Paragraph("Market Overview", styles['Heading2']))
    elements.append(Spacer(1, 8))
    market_data = [["Ticker", "Name", "Tier", "Price", "Change %", "Volume"]]
    for ticker in sorted(current_prices,
                         key=lambda t: current_prices[t].get("change_pct", 0), reverse=True):
        d = current_prices[ticker]
        market_data.append([
            ticker, d.get("name", ""), d.get("tier", ""),
            f"${d.get('price', 0):.2f}", f"{d.get('change_pct', 0):+.2f}%",
            f"{d.get('volume', 0):,}",
        ])
    if len(market_data) > 1:
        elements.append(_styled_table(market_data, '#1a1a2e'))
    else:
        elements.append(Paragraph("No market data available.", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Gainers / Losers
    summary = generate_performance_summary(current_prices, {})
    if "top_gainers" in summary:
        elements.append(Paragraph("Top Gainers", styles['Heading2']))
        g_data = [["Ticker", "Price", "Change %"]]
        for g in summary["top_gainers"]:
            g_data.append([g["ticker"], f"${g['price']:.2f}", f"+{g['change_pct']:.2f}%"])
        widths = [1.5 * inch, 1.5 * inch, 1.5 * inch]
        elements.append(_styled_table(g_data, '#27ae60', widths))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Top Losers", styles['Heading2']))
        l_data = [["Ticker", "Price", "Change %"]]
        for item in summary["top_losers"]:
            l_data.append([item["ticker"], f"${item['price']:.2f}", f"{item['change_pct']:.2f}%"])
        elements.append(_styled_table(l_data, '#e74c3c', widths))
        elements.append(Spacer(1, 20))

    # Holdings
    if holdings:
        elements.append(Paragraph("Portfolio Holdings", styles['Heading2']))
        elements.append(Spacer(1, 8))
        h_data = [["Ticker", "Shares", "Buy Price", "Current", "P&L ($)", "P&L (%)"]]
        for h in holdings:
            ticker = h["ticker"]
            current = current_prices.get(ticker, {}).get("price", h["buy_price"])
            cost = h["shares"] * h["buy_price"]
            pnl = h["shares"] * current - cost
            pnl_pct = (pnl / cost * 100) if cost else 0
            h_data.append([
                ticker, f"{h['shares']:.2f}", f"${h['buy_price']:.2f}",
                f"${current:.2f}", f"${pnl:+.2f}", f"{pnl_pct:+.2f}%",
            ])
        elements.append(_styled_table(h_data, '#2E86C1'))

    doc.build(elements)
    output.seek(0)
    return output
